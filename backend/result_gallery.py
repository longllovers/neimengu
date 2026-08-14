"""Persistent, read-only result gallery for final task artifacts."""

from __future__ import annotations

import csv
import io
import json
import os
import threading
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any


PROJECT_ROOT = Path(__file__).resolve().parents[1]
INDEX_PATH = PROJECT_ROOT / ".geo_workbench_results.json"
LOCK = threading.RLock()
SHP_SIDECARS = (".shp", ".shx", ".dbf", ".prj", ".cpg", ".qix", ".sbn", ".sbx", ".shp.xml")
SUFFIX_KINDS = {
    ".csv": "csv", ".shp": "shp", ".pdf": "pdf", ".json": "json",
    ".geojson": "json", ".xlsx": "excel", ".xls": "excel", ".svg": "svg",
}


def _load() -> list[dict[str, Any]]:
    try:
        data = json.loads(INDEX_PATH.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else []
    except (OSError, json.JSONDecodeError):
        return []


def _save(items: list[dict[str, Any]]) -> None:
    temp = INDEX_PATH.with_suffix(".tmp")
    temp.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temp, INDEX_PATH)


def list_windows() -> list[dict[str, Any]]:
    with LOCK:
        items = _load()
    return [{**item, "assets": [{k: v for k, v in asset.items() if k != "absolute_path"} for asset in item["assets"]]} for item in reversed(items)]


def get_window(window_id: str) -> dict[str, Any]:
    with LOCK:
        item = next((entry for entry in _load() if entry["id"] == window_id), None)
    if item is None:
        raise KeyError(window_id)
    return item


def delete_window(window_id: str) -> None:
    with LOCK:
        items = _load()
        next_items = [item for item in items if item["id"] != window_id]
        if len(items) == len(next_items):
            raise KeyError(window_id)
        _save(next_items)


def delete_all() -> None:
    with LOCK:
        _save([])


def publish_directory(
    tool_id: str, tool_name: str, task_id: str, output_dir: str,
    kinds: set[str] | None = None, newer_than: float | None = None,
) -> dict[str, Any] | None:
    target = Path(output_dir).resolve()
    if not target.exists():
        return None
    root = target if target.is_dir() else target.parent
    allowed_suffixes = {
        suffix for suffix, kind in SUFFIX_KINDS.items()
        if kinds is None or kind in kinds
    }
    source_paths = target.rglob("*") if target.is_dir() else (target,)
    candidates = sorted(
        (
            path for path in source_paths
            if path.is_file() and path.suffix.lower() in allowed_suffixes
            and (newer_than is None or path.stat().st_mtime >= newer_than - 2)
        ),
        key=lambda path: path.as_posix().lower(),
    )[:200]
    if not candidates:
        return None
    assets = []
    for path in candidates:
        kind = SUFFIX_KINDS[path.suffix.lower()]
        assets.append({
            "id": uuid.uuid4().hex, "kind": kind, "name": path.name,
            "relative_path": path.relative_to(root).as_posix(),
            "absolute_path": str(path), "size": path.stat().st_size,
        })
    window = {
        "id": uuid.uuid4().hex, "task_id": task_id, "tool_id": tool_id,
        "title": f"{tool_name} · {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "output_dir": str(root), "created_at": datetime.now().isoformat(timespec="seconds"),
        "assets": assets,
    }
    with LOCK:
        items = [item for item in _load() if item.get("task_id") != task_id]
        items.append(window)
        _save(items[-100:])
    return window


def asset_path(window_id: str, asset_id: str) -> tuple[dict[str, Any], Path]:
    window = get_window(window_id)
    asset = next((entry for entry in window["assets"] if entry["id"] == asset_id), None)
    if asset is None:
        raise KeyError(asset_id)
    path = Path(asset["absolute_path"])
    if not path.is_file():
        raise FileNotFoundError(path)
    return asset, path


def csv_preview(window_id: str, asset_id: str, limit: int = 500) -> dict[str, Any]:
    asset, path = asset_path(window_id, asset_id)
    if asset["kind"] != "csv":
        raise ValueError("不是 CSV 文件")
    rows: list[list[str]] = []
    raw = path.read_bytes()
    try:
        content = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        content = raw.decode("gb18030", errors="replace")
    reader = csv.reader(io.StringIO(content, newline=""))
    for index, row in enumerate(reader):
        if index > limit:
            break
        rows.append([str(value) for value in row])
    headers = rows[0] if rows else []
    return {"name": asset["name"], "headers": headers, "rows": rows[1:], "truncated": len(rows) > limit}


def excel_preview(window_id: str, asset_id: str, limit: int = 500) -> dict[str, Any]:
    asset, path = asset_path(window_id, asset_id)
    if asset["kind"] != "excel":
        raise ValueError("不是 Excel 文件")
    sheets: list[dict[str, Any]] = []
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets[:8]:
                rows = [["" if value is None else str(value) for value in row] for row in sheet.iter_rows(values_only=True, max_row=limit + 2)]
                sheets.append({"name": sheet.title, "headers": rows[0] if rows else [], "rows": rows[1:limit + 1], "truncated": sheet.max_row > limit + 1})
        finally:
            workbook.close()
    else:
        import pandas as pd
        for name, frame in list(pd.read_excel(path, sheet_name=None, nrows=limit + 1).items())[:8]:
            sheets.append({"name": str(name), "headers": [str(value) for value in frame.columns], "rows": [["" if pd.isna(value) else str(value) for value in row] for row in frame.iloc[:limit].itertuples(index=False, name=None)], "truncated": len(frame) > limit})
    return {"name": asset["name"], "sheets": sheets}


def json_preview(window_id: str, asset_id: str, max_chars: int = 300_000) -> dict[str, Any]:
    asset, path = asset_path(window_id, asset_id)
    if asset["kind"] != "json":
        raise ValueError("不是 JSON 文件")
    raw = path.read_bytes()
    try:
        content = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        content = raw.decode("gb18030", errors="replace")
    try:
        content = json.dumps(json.loads(content), ensure_ascii=False, indent=2)
    except json.JSONDecodeError:
        pass
    truncated = len(content) > max_chars
    return {"name": asset["name"], "content": content[:max_chars], "truncated": truncated}


def shapefile_geojson(window_id: str, asset_id: str, max_features: int = 800) -> dict[str, Any]:
    asset, path = asset_path(window_id, asset_id)
    if asset["kind"] != "shp":
        raise ValueError("不是 Shapefile")
    import fiona
    from pyproj import Transformer
    from shapely.geometry import mapping, shape
    from shapely.ops import transform

    features = []
    with fiona.open(path) as source:
        bounds = source.bounds
        span = max(bounds[2] - bounds[0], bounds[3] - bounds[1])
        tolerance = span / 1800 if span else 0
        transformer = None
        if source.crs:
            try:
                transformer = Transformer.from_crs(source.crs, "EPSG:4326", always_xy=True).transform
            except Exception:
                transformer = None
        for index, feature in enumerate(source):
            if index >= max_features or not feature.get("geometry"):
                break
            geometry = shape(feature["geometry"])
            if tolerance:
                geometry = geometry.simplify(tolerance, preserve_topology=True)
            if transformer:
                geometry = transform(transformer, geometry)
            properties = dict(feature.get("properties") or {})
            properties = {str(k): str(v)[:180] for k, v in list(properties.items())[:12]}
            features.append({"type": "Feature", "geometry": mapping(geometry), "properties": properties})
        total = len(source)
    return {"type": "FeatureCollection", "name": asset["name"], "features": features, "total": total, "truncated": total > len(features)}


def archive_paths(window_id: str) -> list[tuple[Path, str]]:
    window = get_window(window_id)
    files: dict[str, Path] = {}
    root = Path(window["output_dir"])
    for asset in window["assets"]:
        path = Path(asset["absolute_path"])
        if asset["kind"] == "shp":
            stem = path.with_suffix("")
            for suffix in SHP_SIDECARS:
                candidate = Path(str(stem) + suffix)
                if candidate.is_file():
                    files[str(candidate.resolve())] = candidate
        elif path.is_file():
            files[str(path.resolve())] = path
    result = []
    for path in files.values():
        try: name = path.relative_to(root).as_posix()
        except ValueError: name = path.name
        result.append((path, name))
    return result
